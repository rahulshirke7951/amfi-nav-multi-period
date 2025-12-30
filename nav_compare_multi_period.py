import os
import json
import pandas as pd
import openpyxl

# ======================================================
# LOAD RULES
# ======================================================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(BASE_DIR, "scheme_rules.json")) as f:
    RULES = json.load(f)

# ======================================================
# HELPERS
# ======================================================

def normalize(s):
    return " ".join(str(s).upper().split())

def clean_text(s):
    for ch in ["-", "–", "—"]:
        s = s.replace(ch, " ")
    return normalize(s)

def extract_base_scheme(name):
    s = clean_text(name)
    for t in RULES["base_scheme_remove_terms"]:
        s = s.replace(t, "")
    return normalize(s)

def exclusion_reason(name):
    for k in RULES["exclusion_rules"]["contains_any"]:
        if k in name.upper():
            return f"Excluded by rule: contains {k}"
    return None

def select_variant(grp):
    for rule in RULES["selection_rules"]["priority_ladder"]:
        m = grp[
            grp["Mutual Fund Name"].str.upper()
            .apply(lambda x: all(k in x for k in rule))
        ]
        if not m.empty:
            return m.iloc[0], grp.drop(m.index)
    return grp.iloc[0], grp.iloc[1:]

# ======================================================
# FLATTEN MERGED CELLS
# ======================================================

def flatten(file, sheet="NAV Data"):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheet]

    for m in list(ws.merged_cells.ranges):
        val = ws.cell(m.min_row, m.min_col).value
        ws.unmerge_cells(str(m))
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                ws.cell(r, c).value = val

    out = file.replace(".xlsx", "_flat.xlsx")
    wb.save(out)
    return out

# ======================================================
# EXTRACT ONE FILE
# ======================================================

def extract(file):
    flat = flatten(file)
    raw = pd.read_excel(flat, header=None).dropna(how="all")

    hdr = raw[
        raw.apply(lambda r: r.astype(str)
        .str.contains("NAV Name", case=False).any(), axis=1)
    ].index[0]

    raw.columns = raw.iloc[hdr]
    df = raw.iloc[hdr + 1:][["NAV Name", "Net Asset Value"]]
    df.columns = ["Mutual Fund Name", "NAV"]

    df["NAV"] = pd.to_numeric(df["NAV"], errors="coerce")
    df = df.dropna()

    total_raw = len(df)

    excluded = []
    eligible = []

    for _, r in df.iterrows():
        reason = exclusion_reason(r["Mutual Fund Name"])
        if reason:
            excluded.append({
                "Mutual Fund Name": r["Mutual Fund Name"],
                "Reason": reason
            })
        else:
            eligible.append(r)

    df = pd.DataFrame(eligible)
    df["Base"] = df["Mutual Fund Name"].apply(extract_base_scheme)
    df["Key"] = df["Mutual Fund Name"].apply(normalize)

    kept = []
    variant_excl = []

    for _, grp in df.groupby("Base"):
        if len(grp) == 1:
            kept.append(grp.iloc[0])
        else:
            keep, drop = select_variant(grp)
            kept.append(keep)
            for _, d in drop.iterrows():
                variant_excl.append({
                    "Mutual Fund Name": d["Mutual Fund Name"],
                    "Reason": "Excluded: variant selection"
                })

    final_df = pd.DataFrame(kept)
    excluded_df = pd.DataFrame(excluded + variant_excl)

    return final_df, excluded_df, total_raw

# ======================================================
# SINGLE COMPARISON RUN
# ======================================================

def run(latest, past, output_file):
    l_df, l_excl, l_raw = extract(latest)
    p_df, p_excl, p_raw = extract(past)

    l_df = l_df.rename(columns={"NAV": "Latest NAV"})
    p_df = p_df.rename(columns={"NAV": "Past NAV"})

    merged = pd.merge(
        l_df,
        p_df[["Key", "Past NAV"]],
        on="Key",
        how="outer",
        indicator=True
    )

    l_nc = merged[merged["_merge"] == "left_only"]
    p_nc = merged[merged["_merge"] == "right_only"]

    comp = merged[merged["_merge"] == "both"].copy()

    l_zero = comp[comp["Latest NAV"] == 0]
    p_zero = comp[comp["Past NAV"] == 0]

    comp = comp[
        (comp["Latest NAV"] != 0) &
        (comp["Past NAV"] != 0)
    ]

    comp["Change"] = comp["Latest NAV"] - comp["Past NAV"]
    comp["Change %"] = (comp["Change"] / comp["Past NAV"] * 100).round(2)

    nav_comp = comp[
        ["Mutual Fund Name", "Latest NAV", "Past NAV", "Change", "Change %"]
    ].sort_values("Change %", ascending=False)

    l_excl = pd.concat([
        l_excl,
        l_zero.assign(Reason="Excluded: zero NAV")[["Mutual Fund Name", "Reason"]],
        l_nc.assign(Reason="Excluded: not comparable")[["Mutual Fund Name", "Reason"]]
    ], ignore_index=True)

    p_excl = pd.concat([
        p_excl,
        p_zero.assign(Reason="Excluded: zero NAV")[["Mutual Fund Name", "Reason"]],
        p_nc.assign(Reason="Excluded: not comparable")[["Mutual Fund Name", "Reason"]]
    ], ignore_index=True)

    rec = pd.DataFrame([
        ["Latest", "Total Raw", l_raw],
        ["Latest", "Included in NAV Comparison", len(nav_comp)],
        ["Latest", "Excluded Schemes", len(l_excl)],

        ["Past", "Total Raw", p_raw],
        ["Past", "Included in NAV Comparison", len(nav_comp)],
        ["Past", "Excluded Schemes", len(p_excl)],
    ], columns=["File Type", "Category", "Count"])

    with pd.ExcelWriter(output_file, engine="openpyxl") as w:
        nav_comp.to_excel(w, "NAV Comparison", index=False)
        l_excl.to_excel(w, "Excluded_Schemes__Latest", index=False)
        p_excl.to_excel(w, "Excluded_Schemes__Past", index=False)
        rec.to_excel(w, "Reconciliation", index=False)

# ======================================================
# MULTI-PERIOD DRIVER
# ======================================================

if __name__ == "__main__":
    data_dir = "data"
    latest_file = os.path.join(data_dir, "Latest.xlsx")

    if not os.path.exists(latest_file):
        raise FileNotFoundError("Latest.xlsx is mandatory.")

    for i in range(1, 5):
        past_file = os.path.join(data_dir, f"Past{i}.xlsx")

        if not os.path.exists(past_file):
            print(f"Skipping Past{i}: file not found")
            continue

        output = f"NAV_Comparison_Latest_vs_Past{i}.xlsx"
        print(f"Running comparison: Latest vs Past{i}")

        run(latest_file, past_file, output)
